"""
Upload poolers guide projections from Excel to Supabase.

Usage:
    pip install openpyxl requests
    python supabase/upload_poolers_guide.py

What this does:
  - For players already in the DB: updates proj_gp, proj_g, proj_a, proj_pts,
    aav, upside, notes, rank, age, team, pos, scouting_category only.
  - For new players: inserts them (last_gp/g/a/pts, tier, risk, etc. left null).
  - Deletes players no longer in the Excel.

Requires these columns to exist (run in Supabase SQL editor if missing):
    ALTER TABLE poolers_players ADD COLUMN IF NOT EXISTS upside int;
    ALTER TABLE poolers_players ADD COLUMN IF NOT EXISTS scouting_category text;
"""

import openpyxl
import requests
import sys
import os
import math

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "guide-poolers", "2026-2027",
    "Guide du pooler.xlsx"
)

SUPABASE_URL = "https://fifurqlitkywtmhgtzeu.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZpZnVycWxpdGt5d3RtaGd0emV1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY3MDIyMjQsImV4cCI6MjA5MjI3ODIyNH0"
    ".KPVPj1qwbSJJMyLR_-AhDcRs0vi2sUU6qbFQ-kH53C0"
)
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
}

# French position abbreviations → English
POS_MAP = {
    "AD": "RW", "AG": "LW",
    "AD/AG": "RW", "AG/AD": "LW",
    "AD/C": "RW", "AG/C": "LW",
    "C/AD": "C",  "C/AG": "C",
    "LW/RW": "LW",
    "C": "C", "LW": "LW", "RW": "RW", "D": "D", "G": "G",
}


def normalize_pos(raw, fallback="F"):
    if not raw:
        return fallback
    key = str(raw).strip().upper()
    return POS_MAP.get(key, key.split("/")[0][:2])


def compute_risk(bandaid, risque):
    b = str(bandaid).strip().upper() if bandaid else ''
    r = str(risque).strip().upper() if risque else ''
    is_injured = b == 'OUI' or b == 'C'
    is_risky   = r == 'OUI'
    is_medium  = r == 'MOYEN'
    if is_risky:
        return 'High'
    if is_injured and is_medium:
        return 'High'
    if is_injured:
        return 'Medium'
    if is_medium:
        return 'Medium'
    return 'Low'


def compute_bust(pos, rank, risk, proj_pts, upside, aav):
    if pos == 'G':
        return False
    pts = proj_pts or 0
    ceiling = upside or pts
    delta = ceiling - pts
    eff = pts / math.sqrt(aav) if aav and aav > 0 else 999
    if risk == 'High' and rank <= 150 and delta <= 20 and eff < 25:
        return True
    return False


def parse_salary(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val / 1_000_000, 3)
    s = str(val).replace("$", "").replace("M", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_players():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    players = []

    # ── FORWARDS ──
    # Cols: 0=Rang, 1=Noms, 2=Tier, 3=Âge, 4=Pos, 5=Équipe, 6=PJ, 7=B, 8=A,
    #       9=Pts, 10=PPG, 11=Salaire, 12=Band-aid, 13=Risqué, 14=Upside, 15=Notes, 16=scouting_category
    ws = wb["Attaquants"]
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if r[0] is None:
            continue
        try:
            int(float(r[0]))
        except (TypeError, ValueError):
            continue
        name = str(r[1]).strip() if r[1] else None
        if not name:
            continue
        players.append({
            "name":              name,
            "age":               int(float(r[3])) if r[3] else None,
            "pos":               normalize_pos(r[4], "F"),
            "team":              str(r[5]).strip() if r[5] else None,
            "proj_gp":           int(float(r[6])) if r[6] else None,
            "proj_g":            int(float(r[7])) if r[7] else None,
            "proj_a":            int(float(r[8])) if r[8] else None,
            "proj_pts":          (int(float(r[7])) if r[7] else 0) + (int(float(r[8])) if r[8] else 0),
            "aav":               parse_salary(r[11]),
            "risk":              compute_risk(r[12], r[13]),
            "upside":            int(float(r[14])) if r[14] else None,
            "notes":             str(r[15]).strip() if r[15] else None,
            "tier":              str(r[2]).strip() if r[2] else None,
            "scouting_category": str(r[16]).strip() if r[16] else None,
        })

    # ── DEFENCEMEN ──
    # Cols: 0=Rang, 1=Noms, 2=Tier, 3=Âge, 4=Pos, 5=Équipe, 6=PJ, 7=B, 8=A,
    #       9=Pts, 10=PPG, 11=Salaire, 12=Band-aid, 13=Risqué, 14=Upside, 15=Notes, 16=scouting_category
    ws = wb["Défenseurs"]
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if r[0] is None:
            continue
        try:
            int(float(r[0]))
        except (TypeError, ValueError):
            continue
        name = str(r[1]).strip() if r[1] else None
        if not name:
            continue
        players.append({
            "name":              name,
            "age":               int(float(r[3])) if r[3] else None,
            "pos":               "D",
            "team":              str(r[5]).strip() if r[5] else None,
            "proj_gp":           int(float(r[6])) if r[6] else None,
            "proj_g":            int(float(r[7])) if r[7] else None,
            "proj_a":            int(float(r[8])) if r[8] else None,
            "proj_pts":          (int(float(r[7])) if r[7] else 0) + (int(float(r[8])) if r[8] else 0),
            "aav":               parse_salary(r[11]),
            "risk":              compute_risk(r[12], r[13]),
            "upside":            int(float(r[14])) if r[14] else None,
            "notes":             str(r[15]).strip() if r[15] else None,
            "tier":              str(r[2]).strip() if r[2] else None,
            "scouting_category": str(r[16]).strip() if r[16] else None,
        })

    # ── GOALIES ──
    # Source: feuille "Gardiens" (source de vérité complète pour l'affichage)
    # Colonnes: 0=Rang, 1=Noms, 2=Tier, 3=Âge, 4=Pos, 5=Équipe,
    #           6=GP, 7=W, 8=L, 9=OTL, 10=SO, 11=W%, 12=Salaire($),
    #           13=Catégorie, 14=Band-aid(OUI), 15=Risqué(Bas/Moyen/Élevé),
    #           16=Upside, 17=Notes
    RISK_FR = {'Bas': 'Low', 'Moyen': 'Medium', 'Élevé': 'High'}
    import json as _json
    ws = wb["Gardiens"]
    for r in list(ws.iter_rows(values_only=True))[1:]:
        name = str(r[1]).strip() if r[1] else None
        if not name:
            continue
        gp   = int(float(r[6]))  if r[6]  is not None else None
        w    = int(float(r[7]))  if r[7]  is not None else None
        l    = int(float(r[8]))  if r[8]  is not None else None
        otl  = int(float(r[9]))  if r[9]  is not None else None
        so   = int(float(r[10])) if r[10] is not None else None
        aav  = parse_salary(r[12])
        cat  = str(r[13]).strip() if r[13] else None
        bandaid = bool(r[14])
        risk_fr = str(r[15]).strip() if r[15] else 'Bas'
        risk = RISK_FR.get(risk_fr, 'Low')
        upside = int(float(r[16])) if r[16] is not None else None
        notes = str(r[17]).strip() if r[17] else None
        goalie_stats = _json.dumps({
            'gp':  gp,
            'w':   w,
            'l':   l,
            'otl': otl,
            'so':  so,
        })
        players.append({
            "name":              name,
            "age":               int(float(r[3])) if r[3] is not None else None,
            "pos":               "G",
            "team":              str(r[5]).strip() if r[5] else "FA",
            "proj_gp":           gp,
            "proj_g":            w,
            "proj_a":            so,
            "proj_pts":          int(float(r[0])) if r[0] is not None else None,
            "aav":               aav,
            "risk":              risk,
            "tier":              str(r[2]).strip() if r[2] else None,
            "notes":             notes,
            "scouting_category": cat,
            "upside":            upside,
            "bandaid":           bandaid,
            "scouting":          goalie_stats,
        })

    # Global rank by proj_pts descending (nulls last), then sheet order as tiebreak.
    players_with_pts = [(p["proj_pts"] or 0, i, p) for i, p in enumerate(players)]
    players_with_pts.sort(key=lambda x: (-x[0], x[1]))
    for rank, (_, _, p) in enumerate(players_with_pts, start=1):
        p["rank"] = rank
        p["bust_alert"] = compute_bust(p["pos"], rank, p["risk"], p["proj_pts"], p.get("upside"), p["aav"])

    return players


def fetch_existing():
    res = requests.get(
        f"{SUPABASE_URL}/rest/v1/poolers_players?select=id,name",
        headers=HEADERS,
    )
    if res.status_code != 200:
        print(f"Failed to fetch existing players: {res.status_code} {res.text}")
        sys.exit(1)
    return {row["name"].lower(): row["id"] for row in res.json()}


def upload(players):
    existing = fetch_existing()
    print(f"  {len(existing)} players already in DB.")

    excel_names = {p["name"].lower() for p in players}
    to_update = []
    to_insert = []
    to_delete = [(name, pid) for name, pid in existing.items() if name not in excel_names]
    for p in players:
        if p["name"].lower() in existing:
            to_update.append((existing[p["name"].lower()], p))
        else:
            to_insert.append(p)

    print(f"  {len(to_update)} to update, {len(to_insert)} to insert, {len(to_delete)} to delete.")

    PROJ_FIELDS = ["rank", "name", "age", "pos", "team",
                   "proj_gp", "proj_g", "proj_a", "proj_pts",
                   "aav", "risk", "tier", "bust_alert", "upside", "notes",
                   "scouting_category", "bandaid"]
    GOALIE_EXTRA = ["scouting"]

    updated = 0
    for pid, p in to_update:
        patch = {k: p.get(k) for k in PROJ_FIELDS}
        if p.get('pos') == 'G':
            for k in GOALIE_EXTRA:
                if k in p:
                    patch[k] = p[k]
        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/poolers_players?id=eq.{pid}",
            headers={**HEADERS, "Prefer": "return=minimal"},
            json=patch,
        )
        if res.status_code not in (200, 204):
            print(f"  PATCH failed for {p['name']}: {res.status_code} {res.text[:200]}")
            sys.exit(1)
        updated += 1
        if updated % 50 == 0:
            print(f"  Updated {updated}/{len(to_update)}…")
    if to_update:
        print(f"  Updated {updated}/{len(to_update)}.")

    if to_insert:
        BATCH = 200
        inserted = 0
        for i in range(0, len(to_insert), BATCH):
            batch = to_insert[i : i + BATCH]
            res = requests.post(
                f"{SUPABASE_URL}/rest/v1/poolers_players",
                headers={**HEADERS, "Prefer": "return=minimal"},
                json=batch,
            )
            if res.status_code not in (200, 201):
                print(f"  INSERT failed at row {i}: {res.status_code} {res.text[:300]}")
                sys.exit(1)
            inserted += len(batch)
            print(f"  Inserted {inserted}/{len(to_insert)}…")

    if to_delete:
        BATCH = 100
        deleted = 0
        delete_ids = [pid for _, pid in to_delete]
        for i in range(0, len(delete_ids), BATCH):
            batch = delete_ids[i:i + BATCH]
            id_list = ",".join(str(x) for x in batch)
            res = requests.delete(
                f"{SUPABASE_URL}/rest/v1/poolers_players?id=in.({id_list})",
                headers=HEADERS,
            )
            if res.status_code not in (200, 204):
                print(f"  DELETE failed: {res.status_code} {res.text[:200]}")
                sys.exit(1)
            deleted += len(batch)
            print(f"  Deleted {deleted}/{len(delete_ids)}…")
        print(f"  Deleted {deleted}/{len(delete_ids)}.")

    print(f"\nDone! {len(to_update)} updated, {len(to_insert)} inserted, {len(to_delete)} deleted.")


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    print("Parsing Excel…")
    players = parse_players()
    f = sum(1 for p in players if p["pos"] not in ("D", "G"))
    d = sum(1 for p in players if p["pos"] == "D")
    g = sum(1 for p in players if p["pos"] == "G")
    cats = sum(1 for p in players if p.get("scouting_category"))
    print(f"  {len(players)} players parsed  ({f} F, {d} D, {g} G)")
    print(f"  {cats} players with scouting_category")
    print()
    print("Sample:")
    for p in players[:3]:
        print(f"  #{p['rank']:3d} {p['name']:<25} {p['pos']:<3} {p['team']:<4}  "
              f"PTS:{p['proj_pts']}  AAV:${p['aav']}M  Cat:{p.get('scouting_category') or '—'}")

    print()
    confirm = input("Upload to Supabase? [y/N] ")
    if confirm.lower() != "y":
        print("Aborted.")
        sys.exit(0)

    upload(players)


if __name__ == "__main__":
    main()
